import time as dt
import openpyxl as op
file1=op.load_workbook("U:\\py\\help1.xlsx")
active_sheet=file1.active
row=2
column=2
name=input('Enter the name')
list_of_names=active_sheet._cells_by_col(2,2,2,10)
for i in list_of_names:
    print(i.value)
while active_sheet.cell(row,column).value != None:
    if name.upper()==active_sheet.cell(row,column).value.upper():
        time1=dt.gmtime()
        active_sheet.cell(row,4).value=dt.strftime("%Y-%m-%d %H:%M:%S",time1)
    

    row+=1
file1.save("U:\\py\\help1.xlsx")